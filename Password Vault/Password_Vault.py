# -*- coding: utf-8 -*-
"""
Created on Sat Jan 23 11:33:12 2021

@author: Tanmoyee
"""

from tkinter import *
import tkinter as tk
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

main=Tk()
main.title("Password Vault")
main.geometry("800x300")
main.config(highlightbackground="black",highlightthickness=2)

file=pathlib.Path("Passwords.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Application Name"
    sheet["B1"]="Application URL"
    sheet["C1"]="Password"
    
    file.save("Passwords.xlsx")
    
def show():
    if var.get()==1:
        passEntry.config(show='')
    if var.get()==0:
        passEntry.config(show='*')
        
def submit():
    y=Application_Name.get()
    z=Application_URL.get()
    z1=passEntry.get()
    print(y)
    print(z)
    print(z1)
    
    file=openpyxl.load_workbook("Passwords.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=y)
    sheet.cell(column=2,row=sheet.max_row,value=z)
    sheet.cell(column=3,row=sheet.max_row,value=z1)
    
    file.save("Passwords.xlsx")
        
frame=LabelFrame(main,text='Password Details').pack(expand='yes',fill='both')
Label(frame,text="Application Name:").place(x=50,y=50)
Label(frame,text="Application URL:").place(x=50,y=90)
Label(frame,text="Password:").place(x=50,y=140)

Application_Name=Entry(frame)
Application_Name.place(x=250,y=50)
Application_URL=Entry(frame)
Application_URL.place(x=250,y=90)
Password=StringVar()
passEntry=Entry(frame,textvariable=Password,show='*')
passEntry.place(x=250,y=140)

var=IntVar()
var.set('0')
chkbtn=Checkbutton(frame,text='Show',variable=var,onvalue=1,offvalue=0,command=show).place(x=350,y=140)
Button(frame,text="Submit",command=submit).place(x=400,y=250)
main.mainloop()